

import os
from flask import Flask, render_template, request
import openpyxl
import pandas as pd
from deps import deps_list 
from flaskwebgui import FlaskUI
import xlsxwriter

app = Flask(__name__)
ui = FlaskUI(app)
list_of_deps = {}

def organize_results(folder_path):
    os.chdir(folder_path)
    files = os.listdir(folder_path)
    
    try:
        os.mkdir("mods")
    except FileExistsError:
        pass
    for f in files:
        if f.endswith('.xlsx'):
            file_path = str(folder_path) + "/{}".format(f)
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            dep = ws['C2'] #Sınavın isminin ve öğrenci bölümünün yazdığı bilgiler hücresi
            
            for k,v in deps_list.items():
                if k in dep.value: #deps.py'da dict olarak verilen listedekilerden(dep_list) birisi eğer C2'de geçiyorsa onu bölüm ismi(dep_name) olarak alır
                    dep_name = v

            ws.delete_cols(2,2) #B:C arası gereksizleri sil
            ws.delete_cols(3,6) #J sil
            ws.delete_rows(1,1) #1.satırı sil

            col_A = []
            col_B = []

            ws.column_dimensions['A'].width = 20 # A sütun genişiliğini azaltma
            for col_cell in ws['A']: # A sütunundaki hücreleri number formatına getirir.
                try:
                    col_cell.number_format = '0'
                    col_cell.value = int(col_cell.value)
                    col_A.append(col_cell.value)
                except ValueError: #Bu kısma öğrenci numarası ALMS'de görünmeyen öğrenciler için bir uyarı eklenecek.
                    pass

            for col_cell in ws['B']: 
                if col_cell.value != None:
                    try:
                        col_cell.number_format = '0'# B sütunundaki hücreleri number formatına getirir.
                        col_cell.value = int(float(col_cell.value.replace(",",".")))#Sayı decimal işaretini ,'den .'ya çevirir.
                    except AttributeError:
                        pass
                else:
                    pass
                col_B.append(col_cell.value)

            all_data = zip(col_A,col_B) #col_A ve col_B'de topladıklarımızı eşleştir.
            all_data_list = list(all_data) #üsttekini listeye çevir

            sorted_list = sorted(all_data_list, key=lambda x: x[1], reverse=True) #Notları büyükten küçüğe sıralar

            df = pd.DataFrame(sorted_list) #Sıralanmış halini pandas dataframe'e çevirir

            writer = pd.ExcelWriter("mods/{}.xlsx".format(dep_name), 
                                    engine='xlsxwriter',
                                    )
            df.to_excel(writer,index=False, sheet_name='Sheet1')
            wb_alldeps = writer.book
            ws_alldeps = writer.sheets['Sheet1']
            fmt = wb_alldeps.add_format({'num_format':'0'}) #Sayı formatını belirle
            ws_alldeps.set_column('A:B', 20, fmt) #Gerekli sütunları seç ve sayı formatını uygula
            writer.save()

def create_data(data,id,list_of_deps):
    mods_path = folder_path + "/mods/"
    
    for k,v in list_of_deps.items():
        if id == v:
            chosen_file = mods_path + v + ".xlsx"
            wb = openpyxl.load_workbook(chosen_file)
            ws = wb.active
            copied_cells = ws['A2:B200']
            for a,b in copied_cells:
                if a.value and b.value != -1:
                    data.append([a.value,b.value])

    df = pd.DataFrame(data=data,columns=["Num","Grade"])
    tr_list = df.drop_duplicates(subset=["Num"]) #ALMS'nin zaman zaman verdiği birden fazla kaydı engellemek için
    tr_list.to_clipboard(excel=True,header=False,index=False) 
    print("************************\n* {}. Copied to clipboard! *\n************************".format(id))

def split_path(folder_path):
    absolute = os.path.abspath(folder_path)
    organize_results(absolute)
    mods_path = folder_path + "/mods/"

def list_files(folder_path): #Mods klasöründe dosyaları oluşturur ve oluşturulan dosyaları listeler
    files = os.listdir(folder_path)
    idx = 1
    
    for f in files:
        if f.endswith('.xlsx'):
            file_path = str(folder_path) + "/{}".format(f)
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            dep = ws['C2'] #Sınavın isminin ve öğrenci bölümünün yazdığı bilgiler hücresi
            
            for k,v in deps_list.items():
                if k in dep.value: #Yukarıda dict olarak verilen listedekilerden(dep_list) birisi eğer C2'de geçiyorsa onu bölüm ismi(dep_name) olarak alır
                    dep_name = v
                    list_of_deps[idx] = dep_name
                    #print(idx,":",dep_name)
                    list_of_deps[idx] = dep_name
                    idx += 1

def choose_file(list_of_deps,folder_path):
    mods_path = folder_path + "/mods/"
    
    data = []
    try:
        choice_int = int(choice)
        try:
            create_data(data,choice_int,list_of_deps)
        except ValueError:
            print("********************************\n* Please choose a valid option *\n********************************")
            choose_file(list_of_deps,folder_path)

        list_files(folder_path) 
    except:
        if choice == "q":
            print("Quitting...choose")
            sys.exit()

        else:
            print("********************************\n* Please choose a valid option *\n********************************")
            list_files(folder_path)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/mods_created', methods=['POST'])
def mods_created():
    global folder_path
    folder_path = request.form['folder_path']
    split_path(folder_path)
    list_files(folder_path)
    
    return render_template('mods_created.html', len=len(list_of_deps)+1,list_of_deps=list_of_deps)

@app.route('/copied/<id>')
def copied(id):
    print(id)
    mods_path = folder_path + "/mods/"
    data = []
    
    create_data(data,id,list_of_deps)

    return render_template('copied.html', len=len(list_of_deps)+1,list_of_deps=list_of_deps,id=id)


if __name__ == '__main__':
    try:
        ui.run()
    except OSError:
        print("Please kill the process using port 5000")